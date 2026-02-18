"""
leer_excel.py — Lee el archivo Excel de levantamiento y guarda resultado en .tmp
"""
import openpyxl
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook(r'levantamiento componentes pañol.xlsx', data_only=True)
output = []

for name in wb.sheetnames:
    ws = wb[name]
    output.append(f"=== HOJA: {name} === (filas={ws.max_row}, cols={ws.max_column})")
    
    for row_idx in range(1, min(20, ws.max_row + 1)):
        vals = []
        for col_idx in range(1, min(ws.max_column + 1, 12)):
            c = ws.cell(row=row_idx, column=col_idx)
            v = str(c.value)[:50] if c.value is not None else "[vacio]"
            vals.append(v)
        output.append(f"  R{row_idx}: {' | '.join(vals)}")
    
    output.append(f"  ... total: {ws.max_row} filas")
    output.append("")

# Also read last few rows
for name in wb.sheetnames:
    ws = wb[name]
    output.append(f"=== ULTIMAS FILAS: {name} ===")
    for row_idx in range(max(1, ws.max_row - 3), ws.max_row + 1):
        vals = []
        for col_idx in range(1, min(ws.max_column + 1, 12)):
            c = ws.cell(row=row_idx, column=col_idx)
            v = str(c.value)[:50] if c.value is not None else "[vacio]"
            vals.append(v)
        output.append(f"  R{row_idx}: {' | '.join(vals)}")
    output.append("")

result = "\n".join(output)
with open(".tmp/excel_structure.txt", "w", encoding="utf-8") as f:
    f.write(result)

print("OK - Resultado guardado en .tmp/excel_structure.txt")
