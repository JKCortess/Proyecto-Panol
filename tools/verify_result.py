import openpyxl

wb = openpyxl.load_workbook('LEVANTamiento ENRIQUECIDO.xlsx')
ws = wb.active

# Show 5 representative samples, one at a time
samples = [
    (4, "Rodamiento rigido"),
    (29, "Rodamiento grande"),
    (72, "Caja/Housing"),
    (79, "Chumacera UCP"),
    (104, "Termoplastico"),
    (112, "Rodamiento YARD"),
    (155, "Inserto UC"),
]

for r, label in samples:
    item = ws.cell(row=r, column=1).value
    marca = ws.cell(row=r, column=3).value
    desc = ws.cell(row=r, column=7).value
    valor = ws.cell(row=r, column=9).value
    print(f"--- Fila {r} ({label}) ---")
    print(f"ITEM: {item} | MARCA: {marca}")
    print(f"DESC: {desc}")
    print(f"VALOR: {valor}")
    print()

# Count filled rows in new columns
filled = 0
for r in range(4, 170):
    if ws.cell(row=r, column=7).value:
        filled += 1

print(f"Total filas con datos en columna G: {filled}")
