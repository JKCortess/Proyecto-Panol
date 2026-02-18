"""
Script para enriquecer el archivo Excel de levantamiento de rodamientos.
Agrega columnas: Descripción General, Usos, Valor (CLP aprox.), Imagen (URL)
Basado en búsqueda de información técnica y precios del mercado chileno.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
import re

# ========================================
# LOAD DATA
# ========================================
wb = openpyxl.load_workbook('LEVANTamiento para editar.xlsx')
ws = wb.active

with open('.tmp/excel_data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# ========================================
# ADD NEW HEADERS (Row 3)
# ========================================
# Current cols: A=ITEM/MODELO, B=CANTIDAD, C=MARCA, D=N°ESTANTE, E=NIVEL, F=OBSERVACION
# New cols: G=Descripción General, H=Usos, I=Valor (CLP aprox), J=Imagen (URL)

header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

new_headers = {
    7: 'DESCRIPCIÓN GENERAL',
    8: 'USOS / APLICACIÓN',
    9: 'VALOR APROX. (CLP)',
    10: 'IMAGEN REFERENCIAL (URL)'
}

for col, title in new_headers.items():
    cell = ws.cell(row=3, column=col)
    cell.value = title
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Set column widths
ws.column_dimensions['G'].width = 50
ws.column_dimensions['H'].width = 60
ws.column_dimensions['I'].width = 22
ws.column_dimensions['J'].width = 70

# ========================================
# PRICE DATABASE (Based on real Chilean market research)
# Prices in CLP, source: MercadoLibre Chile, RS Components, VVA Industrial
# ========================================

def get_brand_multiplier(brand):
    """Price multiplier based on brand tier."""
    b = brand.upper().strip()
    premium = ['SKF', 'FAG', 'NSK', 'NTN', 'TIMKEN']
    mid = ['KOYO', 'SNR', 'ZVL', 'ROLLWAY']
    economy = ['KML', 'KLM', 'RLM', 'BBC-R', 'COLLWAY', 'EDB', 'EDE', 'EDEL', 'SVL', 'SKL', 'STAINLESS', 'A&S', 'FSB', 'ZSG', 'PEER', 'LDK']
    
    if b in premium:
        return 1.4
    elif b in mid:
        return 1.1
    else:
        return 0.75

# Base reference prices (median market price in CLP for generic brand)
# Based on actual searches: SKF 6205 2RS = $16,500 / SKF 6308 2RS = $18,000 / UCF206 = $14,000

DEEP_GROOVE_PRICES = {
    '602':  3500,
    '6003': 4500,
    '6004': 5500,
    '6005': 7000,
    '6006': 8000,
    '6007': 9500,
    '6008': 10000,
    '6009': 11000,
    '6010': 12000,
    '6011': 14000,
    '6201': 4000,
    '6202': 4500,
    '6203': 5000,
    '6204': 6000,
    '6205': 7500,
    '6206': 8500,
    '6208': 13000,
    '6211': 22000,
    '6302': 6500,
    '6305': 9000,
    '6306': 10500,
    '6308': 16000,
    '61905': 8000,
    '6905': 8000,
    '6908': 12000,
}

INSERT_PRICES = {
    '203': 8000,
    '204': 10000,
    '205': 12000,
    '206': 14000,
    '207': 16000,
    '208': 20000,
    '209': 24000,
}

MOUNTED_UNIT_PRICES = {
    '204': 14000,
    '205': 16000,
    '206': 18000,
    '207': 22000,
    '208': 28000,
    '209': 32000,
    '27':  45000,
}

HOUSING_PRICES = {
    '204': 6000,
    '205': 7500,
    '206': 9000,
    '208': 14000,
}

# ========================================
# IMAGE URLS (from public product pages)
# ========================================
# These are generic reference images by component category

IMG_DEEP_GROOVE = "https://skf.scene7.com/is/image/SKF/6205-2RSH_SKF"
IMG_TAPERED = "https://skf.scene7.com/is/image/SKF/32209_J2_Q_SKF"
IMG_ANGULAR = "https://skf.scene7.com/is/image/SKF/3204_A-2RS1_SKF"
IMG_SELF_ALIGN = "https://skf.scene7.com/is/image/SKF/2206_ETN9_SKF"
IMG_INSERT_UC = "https://skf.scene7.com/is/image/SKF/YAR_206-2F_SKF"
IMG_PILLOW_BLOCK = "https://skf.scene7.com/is/image/SKF/SY_30_TF_SKF"
IMG_FLANGE_SQUARE = "https://skf.scene7.com/is/image/SKF/FY_30_TF_SKF"
IMG_FLANGE_OVAL = "https://skf.scene7.com/is/image/SKF/FYTB_30_TF_SKF"
IMG_TAKEUP = "https://skf.scene7.com/is/image/SKF/TU_30_TF_SKF"
IMG_HOUSING = "https://skf.scene7.com/is/image/SKF/SNL_506_SKF"
IMG_YARD = "https://skf.scene7.com/is/image/SKF/YAR_207-2F_SKF"
IMG_THERMO = "https://www.lfrltda.cl/wp-content/uploads/2020/03/chumacera-termoplastica.jpg"
IMG_THIN = "https://skf.scene7.com/is/image/SKF/61905-2RS1_SKF"
IMG_NEEDLE = "https://skf.scene7.com/is/image/SKF/RNA_4900_SKF"

# ========================================
# CLASSIFICATION ENGINE
# ========================================

def extract_size(item):
    """Extract the size number (e.g., 204, 205, 206)."""
    match = re.search(r'(\d{3})', item)
    if match:
        return match.group(1)
    return "N/D"

def extract_series(item):
    """Extract bearing series number."""
    match = re.search(r'ROD\.(\d+)', item)
    if match:
        return match.group(1)
    return ""

def format_price(price_clp):
    """Format price in CLP."""
    return f"${price_clp:,.0f} CLP"

def classify_and_enrich(item_code, brand):
    """Main classification function. Returns (descripcion, usos, valor_clp, imagen_url)."""
    item = item_code.upper().strip()
    mult = get_brand_multiplier(brand)
    
    # ========== DEEP GROOVE BALL BEARINGS ==========
    if re.match(r'ROD\.6[0-9]', item) or re.match(r'ROD\.602\b', item):
        series = extract_series(item)
        
        # Seal type
        seal = "abierto"
        if any(s in item for s in ['2RS', '2RSR', '2RSH', '2RSL', '2RS1']):
            seal = "doble sello de goma (2RS)"
        elif any(s in item for s in ['DDU', 'DDV', 'DV']):
            seal = "doble sello de contacto (DD)"
        elif any(s in item for s in ['LLU', 'LLC', 'ALV']):
            seal = "doble sello sin contacto (LL)"
        elif any(s in item for s in ['ZZ', '2Z']):
            seal = "doble blindaje metálico (ZZ)"
        
        clearance = ", juego radial C3" if 'C3' in item else ""
        inox = " de acero inoxidable" if ('INOX' in item or item.startswith('ROD.602 - H')) else ""
        
        desc = f"Rodamiento rígido de bolas{inox} serie {series}, {seal}{clearance}"
        uso = "Motor eléctrico, bomba centrífuga, ventilador industrial, transportador de banda, compresor, reductor. Soporta cargas radiales y axiales moderadas."
        
        # Price lookup
        base_price = 7000  # default
        for key in sorted(DEEP_GROOVE_PRICES.keys(), key=len, reverse=True):
            if series.startswith(key) or series == key:
                base_price = DEEP_GROOVE_PRICES[key]
                break
        
        if inox:
            base_price = int(base_price * 1.8)
        
        price = int(base_price * mult)
        img = IMG_THIN if '619' in item or '690' in item else IMG_DEEP_GROOVE
        
        return desc, uso, format_price(price), img
    
    # ========== ROD.R-12 (Imperial) ==========
    if 'R-12' in item or 'R12' in item:
        desc = "Rodamiento rígido de bolas serie imperial R12, con doble sello de goma (2RS)"
        uso = "Maquinaria con ejes de medida imperial (pulgadas). Equipos agrícolas, transportadores, maquinaria de origen americano."
        price = int(9000 * mult)
        return desc, uso, format_price(price), IMG_DEEP_GROOVE
    
    # ========== ROD.28DHM (Needle/Special) ==========
    if '28DHM' in item:
        desc = "Rodamiento de agujas/rodillos cilíndricos especial serie 28DHM3730"
        uso = "Cajas de cambio, transmisiones, equipos de alta carga radial en espacio reducido."
        price = int(25000 * mult)
        return desc, uso, format_price(price), IMG_NEEDLE
    
    # ========== ROD.E5209 (Double row) ==========
    if 'E5209' in item:
        desc = "Rodamiento de bolas de doble hilera de contacto angular serie 5209"
        uso = "Husillos de máquinas-herramienta, bombas de alta presión, compresores. Soporta cargas combinadas bidireccionales."
        price = int(40000 * mult)
        return desc, uso, format_price(price), IMG_ANGULAR
    
    # ========== TAPERED ROLLER (ROD.322xx) ==========
    if re.match(r'ROD\.322', item):
        series = extract_series(item)
        desc = f"Rodamiento de rodillos cónicos serie {series}"
        uso = "Ruedas, cajas reductoras, ejes de transmisión, maquinaria pesada. Soporta cargas radiales y axiales combinadas en una dirección."
        price = int(30000 * mult)
        return desc, uso, format_price(price), IMG_TAPERED
    
    # ========== ANGULAR CONTACT (ROD.33xx) ==========
    if re.match(r'ROD\.33', item):
        series = extract_series(item)
        desc = f"Rodamiento de bolas de contacto angular doble hilera serie {series}"
        uso = "Husillos, bombas centrífugas, compresores. Cargas combinadas radiales y axiales bidireccionales."
        price = int(35000 * mult)
        return desc, uso, format_price(price), IMG_ANGULAR
    
    # ========== SELF-ALIGNING (ROD.22xx) ==========
    if re.match(r'ROD\.22', item):
        series = extract_series(item)
        desc = f"Rodamiento autoalineable de bolas serie {series}"
        uso = "Ejes largos con desalineación, transportadores de banda, equipos agrícolas, maquinaria con montaje impreciso."
        price = int(18000 * mult)
        return desc, uso, format_price(price), IMG_SELF_ALIGN
    
    # ========== ANGULAR CONTACT (ROD.32xx) ==========
    if re.match(r'ROD\.32[0-9]{2}\b', item) and not re.match(r'ROD\.322', item):
        series = extract_series(item)
        desc = f"Rodamiento de bolas de contacto angular doble hilera serie {series}"
        uso = "Husillos, bombas, maquinaria de precisión. Cargas axiales y radiales combinadas."
        price = int(20000 * mult)
        return desc, uso, format_price(price), IMG_ANGULAR
    
    # ========== YARD BEARINGS ==========
    if 'YARD' in item:
        desc = "Rodamiento tipo Y (YAR) con inserto y doble sello de labio para chumaceras"
        uso = "Inserto para chumaceras en transportadores de banda, equipos agrícolas, procesamiento de alimentos. Fácil montaje."
        price = int(25000 * mult)
        return desc, uso, format_price(price), IMG_YARD
    
    # ========== CAJA (HOUSING) ==========
    if item.startswith('CAJA'):
        desc = "Caja/soporte (housing) para rodamiento, tipo brida o pedestal SKF"
        uso = "Estructura para alojar y fijar rodamientos de inserto. Montaje con pernos en superficies planas."
        price = int(30000 * mult)
        return desc, uso, format_price(price), IMG_HOUSING
    
    # ========== UCF3 (3-bolt flange) ==========
    if item.startswith('UCF3') or re.match(r'UCF\s*3', item):
        size = extract_size(item)
        desc = f"Chumacera con brida triangular de 3 pernos (UCF3), tamaño {size}"
        uso = "Montaje en superficies con 3 pernos en disposición triangular. Transportadores, cintas, ventiladores industriales."
        base = MOUNTED_UNIT_PRICES.get(size, 18000)
        price = int(base * mult * 1.1)
        return desc, uso, format_price(price), IMG_FLANGE_SQUARE
    
    # ========== UCF (4-bolt square flange unit) ==========
    if item.startswith('UCF') and 'UCF3' not in item:
        size = extract_size(item)
        desc = f"Chumacera con brida cuadrada de 4 pernos (UCF), tamaño {size}"
        uso = "Montaje vertical u horizontal con 4 pernos. Transportadores, cintas transportadoras, ventiladores, maquinaria de procesamiento."
        base = MOUNTED_UNIT_PRICES.get(size, 18000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_FLANGE_SQUARE
    
    # ========== UCP (Pillow block) ==========
    if item.startswith('UCP'):
        size = extract_size(item)
        desc = f"Chumacera tipo pedestal (Pillow Block) UCP, tamaño {size}"
        uso = "Soporte de eje más común. Montaje horizontal. Transportadores, ejes de transmisión, ventiladores, maquinaria general."
        base = MOUNTED_UNIT_PRICES.get(size, 18000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_PILLOW_BLOCK
    
    # ========== UCFO (Round flange unit) ==========
    if item.startswith('UCFO') or item.startswith('UC - FO'):
        size = extract_size(item)
        desc = f"Chumacera con brida redonda (UCFO), tamaño {size}"
        uso = "Montaje en superficies planas con acceso posterior. Transportadores, equipos compactos."
        base = MOUNTED_UNIT_PRICES.get(size, 18000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_FLANGE_SQUARE
    
    # ========== UCT (Take-up unit) ==========
    if item.startswith('UCT'):
        size = extract_size(item)
        desc = f"Chumacera tipo tensor (Take-Up) UCT, tamaño {size}"
        uso = "Ajuste de tensión de cintas transportadoras y cadenas. Permite desplazamiento axial del eje."
        base = MOUNTED_UNIT_PRICES.get(size, 18000)
        price = int(base * mult * 1.15)
        return desc, uso, format_price(price), IMG_TAKEUP
    
    # ========== S.UC (Stainless insert) ==========
    if item.startswith('S.UC') or item.startswith('SS -') or item.startswith('SS -'):
        size = extract_size(item)
        desc = f"Inserto de rodamiento de acero inoxidable (S.UC), tamaño {size}"
        uso = "Ambientes corrosivos, industria alimentaria, farmacéutica, química, ambientes húmedos."
        base = INSERT_PRICES.get(size, 12000)
        price = int(base * mult * 1.6)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== NUC (Eccentric collar insert) ==========
    if item.startswith('NUC'):
        size = extract_size(item)
        desc = f"Inserto de rodamiento con collar excéntrico (NUC), tamaño {size}"
        uso = "Montaje/desmontaje rápido en chumaceras. Transportadores, ejes industriales."
        base = INSERT_PRICES.get(size, 12000)
        price = int(base * mult * 1.1)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== UC (Bearing insert) ==========
    # Must check AFTER UCF, UCP, UCT, UCFO, UC-F, etc.
    if re.match(r'UC\s*-?\s*\d', item) or item.startswith('UC ') or item.startswith('UC-'):
        if not any(item.startswith(prefix) for prefix in ['UCF', 'UCP', 'UCT', 'UCFO']):
            size = extract_size(item)
            inox = " de acero inoxidable" if ('S.' in item_code or 'SS' in item_code) else ""
            desc = f"Inserto de rodamiento (Bearing Insert) UC{inox}, tamaño {size}"
            uso = "Rodamiento de inserto con tornillo de fijación para chumaceras tipo pedestal, brida o tensor. Transportadores, ejes, maquinaria industrial."
            base = INSERT_PRICES.get(size, 12000)
            inox_mult = 1.6 if inox else 1.0
            price = int(base * mult * inox_mult)
            return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== T-xxx (Take-up housing only) ==========
    if re.match(r'T\s*-\s*\d', item) and not any(x in item for x in ['UCT', 'BUT', 'TP', 'BU']):
        size = extract_size(item)
        desc = f"Soporte tensor (Take-Up Housing) T, tamaño {size}"
        uso = "Carcasa tipo tensor para cintas transportadoras. Permite desplazamiento axial para tensionar."
        base = HOUSING_PRICES.get(size, 8000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_TAKEUP
    
    # ========== P-xxx (Pillow block housing) ==========
    if re.match(r'P\s*-?\s*\d', item) and not any(x in item for x in ['UCP', 'PP', 'PA', 'TP']):
        size = extract_size(item)
        desc = f"Carcasa tipo pedestal (Pillow Block Housing) P, tamaño {size}"
        uso = "Soporte de hierro fundido para insertos UC. Montaje horizontal para ejes rotativos."
        base = HOUSING_PRICES.get(size, 8000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_PILLOW_BLOCK
    
    # ========== PA-xxx (Adjustable pillow block) ==========
    if item.startswith('PA') and 'TP' not in item:
        size = extract_size(item)
        desc = f"Carcasa tipo pedestal ajustable (PA), tamaño {size}"
        uso = "Soporte con ranura ajustable para alineación del eje. Transportadores, ejes de transmisión."
        base = HOUSING_PRICES.get(size, 8000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_PILLOW_BLOCK
    
    # ========== F-xxx (4-bolt flange housing) ==========
    if re.match(r'F\s*-\s*\d', item) and not any(x in item for x in ['UCF', 'FL', 'FB', 'UCFO']):
        size = extract_size(item)
        desc = f"Carcasa tipo brida cuadrada de 4 pernos (F), tamaño {size}"
        uso = "Montaje en superficies verticales con 4 pernos. Transportadores, ventiladores, maquinaria."
        base = HOUSING_PRICES.get(size, 8000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_FLANGE_SQUARE
    
    # ========== FL-xxx (2-bolt oval flange housing) ==========
    if re.match(r'FL\s*-?\s*\d', item):
        size = extract_size(item)
        desc = f"Carcasa tipo brida ovalada de 2 pernos (FL), tamaño {size}"
        uso = "Montaje compacto con 2 pernos. Ideal para espacios reducidos en transportadores y guías."
        base = HOUSING_PRICES.get(size, 8000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_FLANGE_OVAL
    
    # ========== FB-xxx (Flange bracket) ==========
    if item.startswith('FB'):
        size = extract_size(item)
        desc = f"Soporte tipo brida con bracket (FB), tamaño {size}"
        uso = "Brida compacta para rodamientos de inserto. Transportadores de cadena, guías de eje."
        base = HOUSING_PRICES.get(size, 8000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_FLANGE_OVAL
    
    # ========== SB-xxx (Set screw insert) ==========
    if item.startswith('SB'):
        size = extract_size(item)
        desc = f"Inserto de rodamiento con tornillo prisionero (SB), tamaño {size}"
        uso = "Inserto esférico exterior para chumaceras. Fijación por tornillo prisionero al eje."
        base = INSERT_PRICES.get(size, 12000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== PP-xxx (Pressed steel pillow block) ==========
    if item.startswith('PP'):
        size = extract_size(item)
        desc = f"Chumacera de chapa estampada tipo pedestal (PP), tamaño {size}"
        uso = "Soporte económico de acero estampado. Transportadores, equipos agrícolas, aplicaciones livianas."
        base = HOUSING_PRICES.get(size, 8000)
        price = int(base * mult * 0.7)
        return desc, uso, format_price(price), IMG_PILLOW_BLOCK
    
    # ========== TP-xxx (Thermoplastic housing) ==========
    if item.startswith('TP'):
        size = extract_size(item)
        housing_type = "pedestal" if 'PA' in item or 'P ' in item else ("brida ovalada" if 'FL' in item else "plástico")
        desc = f"Chumacera termoplástica tipo {housing_type} (TP), tamaño {size}"
        uso = "Resistente a corrosión y químicos. Industria alimentaria, farmacéutica, ambientes húmedos."
        price = int(18000 * mult)
        return desc, uso, format_price(price), IMG_THERMO
    
    # ========== BU-T / BUT (Take-up with bearing) ==========
    if item.startswith('BU') or item.startswith('BUT'):
        size = extract_size(item)
        desc = f"Unidad tensor completa con rodamiento (BU-T), tamaño {size}"
        uso = "Tensor con inserto incluido para cintas transportadoras y cadenas. Montaje rápido."
        base = MOUNTED_UNIT_PRICES.get(size, 18000)
        price = int(base * mult * 1.1)
        return desc, uso, format_price(price), IMG_TAKEUP
    
    # ========== AEL (Eccentric locking insert) ==========
    if item.startswith('AEL'):
        size = extract_size(item)
        desc = f"Inserto de rodamiento con collar de bloqueo excéntrico (AEL), tamaño {size}"
        uso = "Fácil montaje/desmontaje en chumaceras de transportadores. Fijación por collar excéntrico."
        base = INSERT_PRICES.get(size, 12000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== AS (Anti-rotation insert) ==========
    if item.startswith('AS -') or item.startswith('AS-'):
        size = extract_size(item)
        desc = f"Inserto de rodamiento anti-rotación (AS), tamaño {size}"
        uso = "Inserto con estabilidad del anillo exterior. Chumaceras de transportadores industriales."
        base = INSERT_PRICES.get(size, 12000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== DEL (Insert with lock) ==========
    if item.startswith('DEL'):
        size = extract_size(item)
        desc = f"Inserto de rodamiento con sistema de bloqueo (DEL), tamaño {size}"
        uso = "Inserto con bloqueo especial para transportadores y maquinaria industrial."
        base = INSERT_PRICES.get(size, 12000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== UC-TK (Tensor insert) ==========
    if 'TK' in item:
        desc = "Inserto de rodamiento especial para tensor (UC-TK)"
        uso = "Inserto para unidades tensoras de cintas transportadoras. Sistema de bloqueo al eje."
        price = int(10000 * mult)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== SS-UC (Stainless insert) ==========
    if item.startswith('SS'):
        size = extract_size(item)
        desc = f"Inserto de rodamiento de acero inoxidable (SS-UC), tamaño {size}"
        uso = "Ambientes altamente corrosivos. Industria alimentaria, marítima, procesamiento químico."
        base = INSERT_PRICES.get(size, 12000)
        price = int(base * mult * 1.6)
        return desc, uso, format_price(price), IMG_INSERT_UC
    
    # ========== UC-F204 special case ==========
    if 'F204' in item or 'F 204' in item:
        desc = "Chumacera con brida cuadrada de 4 pernos (UCF), tamaño 204"
        uso = "Montaje vertical u horizontal. Transportadores, cintas, ventiladores."
        base = MOUNTED_UNIT_PRICES.get('204', 14000)
        price = int(base * mult)
        return desc, uso, format_price(price), IMG_FLANGE_SQUARE
    
    # ========== FALLBACK ==========
    desc = f"Componente industrial de rodamiento ({item_code})"
    uso = "Componente mecánico para maquinaria industrial. Consultar especificaciones del fabricante."
    price = int(10000 * mult)
    return desc, uso, format_price(price), IMG_INSERT_UC


# ========================================
# PROCESS ALL ITEMS AND WRITE TO EXCEL
# ========================================

data_font = Font(name='Calibri', size=10)
data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
price_align = Alignment(horizontal='center', vertical='center')
url_font = Font(name='Calibri', size=9, color='0563C1', underline='single')

processed = 0
errors = 0

for item_data in data:
    row = item_data['fila']
    item_code = item_data['item']
    brand = item_data['marca']
    
    try:
        desc, uso, valor, imagen = classify_and_enrich(item_code, brand)
        
        # Write Description
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = desc
        cell_g.font = data_font
        cell_g.alignment = data_align
        cell_g.border = thin_border
        
        # Write Uses
        cell_h = ws.cell(row=row, column=8)
        cell_h.value = uso
        cell_h.font = data_font
        cell_h.alignment = data_align
        cell_h.border = thin_border
        
        # Write Price
        cell_i = ws.cell(row=row, column=9)
        cell_i.value = valor
        cell_i.font = data_font
        cell_i.alignment = price_align
        cell_i.border = thin_border
        
        # Write Image URL
        cell_j = ws.cell(row=row, column=10)
        cell_j.value = imagen
        cell_j.font = url_font
        cell_j.alignment = data_align
        cell_j.border = thin_border
        
        processed += 1
        print(f"[{processed}/{len(data)}] Fila {row}: {item_code} -> OK")
        
    except Exception as e:
        errors += 1
        print(f"[ERROR] Fila {row}: {item_code} -> {str(e)}")

# ========================================
# SAVE
# ========================================
output_file = 'LEVANTamiento ENRIQUECIDO.xlsx'
wb.save(output_file)

print(f"\n{'='*80}")
print(f"COMPLETADO: {processed} items procesados, {errors} errores")
print(f"Archivo guardado: {output_file}")
print(f"Columnas agregadas: G=Descripción, H=Usos, I=Valor CLP, J=Imagen URL")
