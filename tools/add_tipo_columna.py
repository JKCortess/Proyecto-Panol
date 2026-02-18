"""
Agrega columna "TIPO DE COMPONENTE" al Excel enriquecido.
Clasifica cada item según su código en una categoría simple.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re

wb = openpyxl.load_workbook('LEVANTamiento ENRIQUECIDO.xlsx')
ws = wb.active

# ========================================
# STYLES
# ========================================
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
data_font = Font(name='Calibri', size=10, bold=True)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ========================================
# SHIFT existing columns G-J → H-K, insert new G = TIPO
# We'll insert the new column BEFORE the current G (Descripción)
# So: move J→K, I→J, H→I, G→H, then write G
# ========================================

# First, shift data from cols 10→11, 9→10, 8→9, 7→8 for ALL rows
max_row = ws.max_row
for row in range(3, max_row + 1):
    for col_from, col_to in [(10, 11), (9, 10), (8, 9), (7, 8)]:
        src = ws.cell(row=row, column=col_from)
        dst = ws.cell(row=row, column=col_to)
        dst.value = src.value
        dst.font = src.font.copy() if src.font else data_font
        dst.alignment = src.alignment.copy() if src.alignment else data_align
        dst.border = src.border.copy() if src.border else thin_border
        dst.fill = src.fill.copy() if src.fill else PatternFill()

# Set header for new column G
cell_g = ws.cell(row=3, column=7)
cell_g.value = 'TIPO DE COMPONENTE'
cell_g.font = header_font
cell_g.fill = header_fill
cell_g.alignment = header_align
cell_g.border = thin_border

# Set column widths
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 50
ws.column_dimensions['I'].width = 60
ws.column_dimensions['J'].width = 22
ws.column_dimensions['K'].width = 70

# ========================================
# CLASSIFICATION: Item code → Tipo de componente
# ========================================

def get_tipo(item_code):
    """Return the component type based on item code."""
    item = item_code.upper().strip()
    
    # --- Deep groove ball bearings (most common) ---
    if re.match(r'ROD\.6[0-9]', item) or re.match(r'ROD\.602\b', item):
        if 'INOX' in item or item.startswith('ROD.602 - H'):
            return 'Rodamiento (Inoxidable)'
        return 'Rodamiento rígido de bolas'
    
    # --- Thin section bearings ---
    if re.match(r'ROD\.619', item) or re.match(r'ROD\.69', item):
        return 'Rodamiento de sección delgada'
    
    # --- Imperial bearing ---
    if 'R-12' in item or 'R12' in item:
        return 'Rodamiento imperial'
    
    # --- Needle / special ---
    if '28DHM' in item:
        return 'Rodamiento de agujas'
    
    # --- Double row angular ---
    if 'E5209' in item:
        return 'Rodamiento contacto angular'
    
    # --- Tapered roller ---
    if re.match(r'ROD\.322', item):
        return 'Rodamiento de rodillos cónicos'
    
    # --- Angular contact ---
    if re.match(r'ROD\.33', item) or re.match(r'ROD\.32[0-9]{2}', item):
        return 'Rodamiento contacto angular'
    
    # --- Self-aligning ---
    if re.match(r'ROD\.22', item):
        return 'Rodamiento autoalineable'
    
    # --- YARD bearings ---
    if 'YARD' in item:
        return 'Rodamiento tipo Y (YAR)'
    
    # --- Housing / Caja ---
    if item.startswith('CAJA'):
        return 'Caja / Housing'
    
    # --- UCF3 (3-bolt flange) ---
    if item.startswith('UCF3') or re.match(r'UCF\s*3', item):
        return 'Chumacera brida 3 pernos'
    
    # --- UCF (4-bolt square flange unit) ---
    if item.startswith('UCF'):
        return 'Chumacera brida cuadrada'
    
    # --- UCP (Pillow block) ---
    if item.startswith('UCP'):
        return 'Chumacera tipo pedestal'
    
    # --- UCFO (Round flange) ---
    if item.startswith('UCFO') or item.startswith('UC - FO'):
        return 'Chumacera brida redonda'
    
    # --- UCT (Take-up) ---
    if item.startswith('UCT'):
        return 'Chumacera tipo tensor'
    
    # --- S.UC / SS-UC (Stainless insert) ---
    if item.startswith('S.UC') or item.startswith('SS -') or item.startswith('SS -'):
        return 'Inserto inoxidable'
    
    # --- NUC (Eccentric collar insert) ---
    if item.startswith('NUC'):
        return 'Inserto con collar excéntrico'
    
    # --- UC (Bearing insert) - after UCF/UCP/UCT ---
    if re.match(r'UC\s*-?\s*\d', item) or item.startswith('UC ') or item.startswith('UC-'):
        if not any(item.startswith(p) for p in ['UCF', 'UCP', 'UCT', 'UCFO']):
            return 'Inserto de rodamiento'
    
    # --- T (Take-up housing) ---
    if re.match(r'T\s*-\s*\d', item) and not any(x in item for x in ['UCT', 'BUT', 'TP', 'BU']):
        return 'Soporte tensor (housing)'
    
    # --- P (Pillow block housing) ---
    if re.match(r'P\s*-?\s*\d', item) and not any(x in item for x in ['UCP', 'PP', 'PA', 'TP']):
        return 'Carcasa tipo pedestal'
    
    # --- PA (Adjustable pillow block) ---
    if item.startswith('PA') and 'TP' not in item:
        return 'Carcasa pedestal ajustable'
    
    # --- F (4-bolt flange housing) ---
    if re.match(r'F\s*-\s*\d', item) and not any(x in item for x in ['UCF', 'FL', 'FB', 'UCFO']):
        return 'Carcasa brida cuadrada'
    
    # --- FL (2-bolt oval flange) ---
    if re.match(r'FL\s*-?\s*\d', item):
        return 'Carcasa brida ovalada'
    
    # --- FB (Flange bracket) ---
    if item.startswith('FB'):
        return 'Soporte brida (bracket)'
    
    # --- SB (Set screw insert) ---
    if item.startswith('SB'):
        return 'Inserto con tornillo prisionero'
    
    # --- PP (Pressed steel pillow block) ---
    if item.startswith('PP'):
        return 'Chumacera chapa estampada'
    
    # --- TP (Thermoplastic housing) ---
    if item.startswith('TP'):
        return 'Chumacera termoplástica'
    
    # --- BU-T / BUT (Take-up with bearing) ---
    if item.startswith('BU') or item.startswith('BUT'):
        return 'Unidad tensor completa'
    
    # --- AEL (Eccentric locking insert) ---
    if item.startswith('AEL'):
        return 'Inserto bloqueo excéntrico'
    
    # --- AS (Anti-rotation insert) ---
    if item.startswith('AS -') or item.startswith('AS-'):
        return 'Inserto anti-rotación'
    
    # --- DEL (Insert with lock) ---
    if item.startswith('DEL'):
        return 'Inserto con bloqueo'
    
    # --- UC-TK ---
    if 'TK' in item:
        return 'Inserto para tensor'
    
    # --- SS-UC ---
    if item.startswith('SS'):
        return 'Inserto inoxidable'
    
    # --- UC-F204 special ---
    if 'F204' in item or 'F 204' in item:
        return 'Chumacera brida cuadrada'
    
    return 'Componente industrial'


# ========================================
# FILL NEW COLUMN G FOR EACH DATA ROW
# ========================================

# Color coding by category type
COLORS = {
    'Rodamiento': PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid'),   # Light blue
    'Chumacera':  PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),    # Light green
    'Inserto':    PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid'),    # Light orange
    'Carcasa':    PatternFill(start_color='F5B7B1', end_color='F5B7B1', fill_type='solid'),    # Light red
    'Soporte':    PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid'),    # Light purple
    'Unidad':     PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),    # Light green
    'Caja':       PatternFill(start_color='F5B7B1', end_color='F5B7B1', fill_type='solid'),    # Light red
    'default':    PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid'),    # Light gray
}

def get_color(tipo):
    for keyword, fill in COLORS.items():
        if keyword.lower() in tipo.lower():
            return fill
    return COLORS['default']

processed = 0
for row in range(4, max_row + 1):
    item_code = ws.cell(row=row, column=1).value
    if item_code is None:
        continue
    
    tipo = get_tipo(str(item_code).strip())
    
    cell = ws.cell(row=row, column=7)
    cell.value = tipo
    cell.font = data_font
    cell.alignment = data_align
    cell.border = thin_border
    cell.fill = get_color(tipo)
    
    processed += 1
    print(f"[{processed}] Fila {row}: {item_code} → {tipo}")

# ========================================
# SAVE
# ========================================
wb.save('LEVANTamiento ENRIQUECIDO.xlsx')

print(f"\n{'='*60}")
print(f"✅ COMPLETADO: {processed} items con tipo de componente")
print(f"Columna G = TIPO DE COMPONENTE (con color por categoría)")
print(f"Archivo: LEVANTamiento ENRIQUECIDO.xlsx")
