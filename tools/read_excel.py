import openpyxl
import json

wb = openpyxl.load_workbook('LEVANTamiento para editar.xlsx')
ws = wb.active

items = []
for r in range(4, ws.max_row + 1):
    item_val = ws.cell(row=r, column=1).value
    cant = ws.cell(row=r, column=2).value
    marca = ws.cell(row=r, column=3).value
    estante = ws.cell(row=r, column=4).value
    nivel = ws.cell(row=r, column=5).value
    obs = ws.cell(row=r, column=6).value
    if item_val is not None:
        items.append({
            'fila': r,
            'item': str(item_val).strip(),
            'cantidad': cant,
            'marca': str(marca).strip() if marca else '',
            'estante': estante,
            'nivel': nivel,
            'observacion': str(obs).strip() if obs else ''
        })

print(f"Total items con datos: {len(items)}")
print("=" * 80)
for i, it in enumerate(items):
    print(f"{i+1}. Fila {it['fila']}: ITEM=[{it['item']}] | MARCA=[{it['marca']}] | CANT={it['cantidad']} | OBS=[{it['observacion']}]")

# Save as JSON for later use
with open('.tmp/excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(items, f, ensure_ascii=False, indent=2)
print("\nDatos guardados en .tmp/excel_data.json")
