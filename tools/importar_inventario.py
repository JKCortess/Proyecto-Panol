"""
importar_inventario.py — Importa datos del Excel al Google Sheet Pañol_DB.

1. Lee 'levantamiento componentes pañol.xlsx'
2. Genera SKUs únicos para cada ítem
3. Mapea ubicaciones (Estante → formato P1-E{estante}-N{nivel})
4. Clasifica ítems y los sube a la hoja ITEMS
5. Actualiza USUARIOS con los emails reales
"""

import openpyxl
import json
import os
import sys
import warnings
import re
warnings.filterwarnings("ignore")

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

# ─── Config ───────────────────────────────────────────────────
TOKEN_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "token.json")
ENV_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "levantamiento componentes pañol.xlsx")
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ─── Usuarios Reales ──────────────────────────────────────────
USUARIOS_REALES = [
    # ID, Nombre, Rol, Email
    ["USR-001", "Administrador Principal", "Admin", ""],  # ← Se llenará con el email del usuario
    ["USR-002", "Daniel Rojas", "Técnico", "daniel.rojas@dole.com"],
    ["USR-003", "Técnico Reserva", "Técnico", ""],
    ["USR-004", "Planificador", "Planificador", ""],
    ["USR-005", "Auditor", "Auditor", ""],
]


def get_spreadsheet_id():
    """Read spreadsheet ID from .env."""
    with open(ENV_PATH, "r") as f:
        for line in f:
            if line.strip().startswith("PANOL_DB_SPREADSHEET_ID="):
                return line.strip().split("=", 1)[1]
    print("❌ No se encontró PANOL_DB_SPREADSHEET_ID en .env")
    sys.exit(1)


def get_credentials():
    """Load OAuth credentials."""
    with open(TOKEN_PATH, "r") as f:
        token_data = json.load(f)
    creds = Credentials(
        token=token_data["token"],
        refresh_token=token_data["refresh_token"],
        token_uri=token_data["token_uri"],
        client_id=token_data["client_id"],
        client_secret=token_data["client_secret"],
        scopes=token_data.get("scopes", SCOPES),
    )
    if creds.expired or not creds.valid:
        creds.refresh(Request())
        token_data["token"] = creds.token
        token_data["expiry"] = creds.expiry.isoformat() + "Z" if creds.expiry else None
        with open(TOKEN_PATH, "w") as f:
            json.dump(token_data, f)
    return creds


def read_excel():
    """Read items from the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    items = []
    sku_counter = {}  # Para manejar duplicados
    
    for row_idx in range(4, ws.max_row + 1):  # Datos empiezan en fila 4
        nombre = ws.cell(row=row_idx, column=1).value
        cantidad = ws.cell(row=row_idx, column=2).value
        marca = ws.cell(row=row_idx, column=3).value
        estante = ws.cell(row=row_idx, column=4).value
        nivel = ws.cell(row=row_idx, column=5).value
        observacion = ws.cell(row=row_idx, column=6).value
        
        # Skip empty rows
        if not nombre or str(nombre).strip() == "":
            continue
        
        nombre = str(nombre).strip()
        
        # Generate SKU from item name
        sku_base = generar_sku(nombre)
        
        # Handle duplicates
        if sku_base in sku_counter:
            sku_counter[sku_base] += 1
            sku = f"{sku_base}-{sku_counter[sku_base]:02d}"
        else:
            sku_counter[sku_base] = 1
            sku = sku_base
        
        # Format ubicacion
        estante_num = str(int(estante)) if estante else "000"
        nivel_num = str(int(nivel)) if nivel else "00"
        ubicacion = f"P1-E{estante_num}-N{nivel_num}"
        
        # Map to category
        categoria = clasificar_item(nombre)
        
        # Determine criticality
        clasificacion = determinar_criticidad(nombre, int(cantidad) if cantidad else 0)
        
        items.append({
            "SKU": sku,
            "Nombre": nombre,
            "Categoria": categoria,
            "Stock_Actual": int(cantidad) if cantidad else 0,
            "Stock_Reservado": 0,
            "Ubicacion": ubicacion,
            "Link_Foto": "",
            "Clasificacion": clasificacion,
            "ROP": 0,  # Se calculará después con historial
            "Safety_Stock": 0,  # Se calculará después
            "Marca": str(marca).strip() if marca else "",
            "Observacion": str(observacion).strip() if observacion else "",
        })
    
    return items


def generar_sku(nombre):
    """Generate a standardized SKU from item name."""
    # Clean the name
    sku = nombre.upper().strip()
    
    # Remove common prefixes
    sku = re.sub(r'^ROD\.', 'ROD-', sku)
    
    # Replace spaces and special chars
    sku = re.sub(r'[\s/]+', '-', sku)
    sku = re.sub(r'[^\w\-]', '', sku)
    
    # Collapse multiple dashes
    sku = re.sub(r'-+', '-', sku)
    sku = sku.strip('-')
    
    # Limit length
    if len(sku) > 30:
        sku = sku[:30].rstrip('-')
    
    return sku


def clasificar_item(nombre):
    """Classify item into a category based on name patterns."""
    nombre_upper = nombre.upper()
    
    if "ROD." in nombre_upper or "RODAMIENTO" in nombre_upper:
        return "Rodamientos"
    elif "CORREA" in nombre_upper or "FAJA" in nombre_upper:
        return "Correas y Fajas"
    elif "CADENA" in nombre_upper:
        return "Cadenas"
    elif "PIÑON" in nombre_upper or "PIÑÓN" in nombre_upper:
        return "Piñones"
    elif "RETENEDOR" in nombre_upper or "SELLO" in nombre_upper or "RETEN" in nombre_upper:
        return "Sellos y Retenes"
    elif "VALVULA" in nombre_upper or "VÁLVULA" in nombre_upper:
        return "Válvulas"
    elif "MOTOR" in nombre_upper:
        return "Motores"
    elif "FILTRO" in nombre_upper:
        return "Filtros"
    elif "TORNILLO" in nombre_upper or "PERNO" in nombre_upper or "TUERCA" in nombre_upper:
        return "Tornillería"
    elif "UC" in nombre_upper or "UCF" in nombre_upper or "UCT" in nombre_upper:
        return "Chumaceras"
    elif "DEL" in nombre_upper:
        return "Rodamientos"  # DEL series are bearings
    else:
        return "General"


def determinar_criticidad(nombre, cantidad):
    """Determine if item is Critical or Non-Critical."""
    # Low stock = more critical
    if cantidad <= 2:
        return "Crítico"
    
    nombre_upper = nombre.upper()
    # Certain components are always critical
    if any(kw in nombre_upper for kw in ["MOTOR", "VALVULA", "VÁLVULA", "ELECTROVALVULA"]):
        return "Crítico"
    
    return "No Crítico"


def upload_items(service, spreadsheet_id, items):
    """Upload items to the ITEMS sheet."""
    # Clear existing data (keep headers)
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range="'ITEMS'!A2:J10000",
    ).execute()
    
    # Prepare rows
    rows = []
    for item in items:
        rows.append([
            item["SKU"],
            item["Nombre"],
            item["Categoria"],
            item["Stock_Actual"],
            item["Stock_Reservado"],
            item["Ubicacion"],
            item["Link_Foto"],
            item["Clasificacion"],
            item["ROP"],
            item["Safety_Stock"],
        ])
    
    body = {
        "valueInputOption": "USER_ENTERED",
        "data": [
            {
                "range": "'ITEMS'!A2",
                "majorDimension": "ROWS",
                "values": rows,
            }
        ],
    }
    
    result = service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id, body=body
    ).execute()
    
    return len(rows)


def update_usuarios(service, spreadsheet_id):
    """Update USUARIOS sheet with real user data."""
    # Clear existing
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range="'USUARIOS'!A2:D100",
    ).execute()
    
    body = {
        "valueInputOption": "USER_ENTERED",
        "data": [
            {
                "range": "'USUARIOS'!A2",
                "majorDimension": "ROWS",
                "values": USUARIOS_REALES,
            }
        ],
    }
    
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id, body=body
    ).execute()
    
    return len(USUARIOS_REALES)


def print_summary(items):
    """Print import summary."""
    categories = {}
    critical_count = 0
    total_stock = 0
    
    for item in items:
        cat = item["Categoria"]
        categories[cat] = categories.get(cat, 0) + 1
        if item["Clasificacion"] == "Crítico":
            critical_count += 1
        total_stock += item["Stock_Actual"]
    
    print(f"\n📊 Resumen de importación:")
    print(f"   Total ítems: {len(items)}")
    print(f"   Stock total: {total_stock} unidades")
    print(f"   Ítems críticos: {critical_count}")
    print(f"\n   Categorías:")
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        print(f"     {cat}: {count} ítems")
    
    print(f"\n   Ubicaciones:")
    ubicaciones = set(item["Ubicacion"] for item in items)
    for ub in sorted(ubicaciones):
        count = sum(1 for i in items if i["Ubicacion"] == ub)
        print(f"     {ub}: {count} ítems")


def main():
    print("=" * 60)
    print("  📥 Importando inventario → Pañol_DB")
    print("=" * 60)
    
    # 1. Read Excel
    print("\n📂 Leyendo Excel...")
    items = read_excel()
    print(f"   ✅ {len(items)} ítems leídos")
    
    # 2. Auth
    creds = get_credentials()
    service = build("sheets", "v4", credentials=creds)
    spreadsheet_id = get_spreadsheet_id()
    print(f"\n🔗 Spreadsheet: {spreadsheet_id}")
    
    # 3. Upload items
    print("\n📤 Subiendo ítems a hoja ITEMS...")
    count = upload_items(service, spreadsheet_id, items)
    print(f"   ✅ {count} ítems subidos")
    
    # 4. Update users
    print("\n👥 Actualizando usuarios...")
    user_count = update_usuarios(service, spreadsheet_id)
    print(f"   ✅ {user_count} usuarios actualizados")
    print(f"   📧 Daniel Rojas (daniel.rojas@dole.com) → Técnico")
    
    # 5. Summary
    print_summary(items)
    
    # 6. Save import log
    log_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".tmp", "import_log.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(f"Importación: {len(items)} ítems\n")
        for item in items:
            f.write(f"  {item['SKU']}: {item['Nombre']} (x{item['Stock_Actual']}) [{item['Categoria']}] @ {item['Ubicacion']}\n")
    print(f"\n📝 Log guardado en .tmp/import_log.txt")
    
    print()
    print("=" * 60)
    print(f"  🎉 Importación completada — {count} ítems en Pañol_DB")
    print(f"  📊 https://docs.google.com/spreadsheets/d/{spreadsheet_id}")
    print("=" * 60)


if __name__ == "__main__":
    main()
